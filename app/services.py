import base64
import html
import json
import math
import re
import xml.etree.ElementTree as ET
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote_plus, urlparse

import httpx
from openpyxl import load_workbook

from app.config import Settings
from app.schemas import Offer, Source


class ServiceConfigurationError(RuntimeError):
    pass


class UpstreamServiceError(RuntimeError):
    pass


def _plain_text(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    return " ".join(html.unescape(value).split())


class _OfferHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []
        self.url = ""

    def handle_data(self, data: str) -> None:
        self.parts.append(data)

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "a" and not self.url:
            href = dict(attrs).get("href", "").strip()
            parsed = urlparse(href)
            if parsed.scheme in {"http", "https"} and parsed.netloc:
                self.url = href


def _clean_offer_html(value: str) -> Tuple[str, str]:
    parser = _OfferHTMLParser()
    try:
        parser.feed(value or "")
        parser.close()
    except Exception:
        return _plain_text(value), ""
    text = "".join(parser.parts).replace("_x000D_", " ")
    return " ".join(text.split()), parser.url


def _tokens(value: str) -> List[str]:
    words = re.findall(r"[a-zа-яё0-9]+", (value or "").lower())
    result = []
    endings = (
        "иями", "ами", "ями", "ого", "ему", "ому", "ыми", "ими",
        "ая", "яя", "ое", "ее", "ие", "ые", "ий", "ый", "ой",
        "ам", "ям", "ах", "ях", "ов", "ев", "ом", "ем", "ами",
        "ы", "и", "а", "я", "у", "ю", "е", "о",
    )
    for word in words:
        normalized = word
        if len(word) > 5:
            for ending in endings:
                if word.endswith(ending) and len(word) - len(ending) >= 4:
                    normalized = word[:-len(ending)]
                    break
        result.append(normalized)
    return result


_SEARCH_STOP_WORDS = {
    "в", "во", "для", "до", "из", "и", "или", "к", "как", "на", "не",
    "но", "о", "от", "по", "под", "при", "руб", "рублей", "с", "со", "у",
}

_RELATED_OFFER_TERMS = {
    "наушник": ("аудио", "электроник", "гаджет", "техник"),
    "подарок": ("подар", "сертификат", "сувенир", "цвет"),
    "подарк": ("подар", "сертификат", "сувенир", "цвет"),
    "ресторан": ("кафе", "бар", "кухн"),
    "тур": ("отел", "путешеств", "билет"),
}


class OfferCatalog:
    """In-memory offer index backed by data/ofrs_merge.xlsx."""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.rows = []
        self.document_frequency: Counter = Counter()
        self.average_length = 1.0
        self._load()

    def _load(self) -> None:
        if not self.workbook_path.exists():
            return
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(iterator)]
        columns = {name: index for index, name in enumerate(headers)}
        required = {"offer_id", "name", "data.description"}
        if not required.issubset(columns):
            workbook.close()
            raise ServiceConfigurationError(
                "В data/ofrs_merge.xlsx отсутствуют обязательные столбцы: "
                + ", ".join(sorted(required - set(columns)))
            )

        lengths = []
        for values in iterator:
            description_html = str(values[columns["data.description"]] or "")
            description, embedded_url = _clean_offer_html(description_html)
            if not description:
                continue
            name = str(values[columns["name"]] or "Предложение").strip()
            offer_id = str(values[columns["offer_id"]] or "")
            image_url = self._cell(values, columns, "source_image")
            categories = self._cell(values, columns, "categories")
            options = self._cell(values, columns, "data.options")
            searchable = " ".join((name, name, categories, options, description))
            title_counts = Counter(_tokens(" ".join((name, categories))))
            counts = Counter(_tokens(searchable))
            if not counts:
                continue
            self.document_frequency.update(counts.keys())
            lengths.append(sum(counts.values()))
            self.rows.append(
                {
                    "id": offer_id,
                    "name": name,
                    "description": description,
                    "url": embedded_url
                    or "https://yandex.ru/search/?text=" + quote_plus(name),
                    "image_url": image_url if self._is_http_url(image_url) else "",
                    "counts": counts,
                    "title_counts": title_counts,
                    "length": sum(counts.values()),
                }
            )
        workbook.close()
        if lengths:
            self.average_length = sum(lengths) / len(lengths)

    @staticmethod
    def _cell(values, columns: Dict[str, int], name: str) -> str:
        index = columns.get(name)
        return str(values[index] or "").strip() if index is not None else ""

    @staticmethod
    def _is_http_url(value: str) -> bool:
        parsed = urlparse(value)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)

    def search(self, query: str, limit: int = 3) -> List[Offer]:
        primary_terms = [
            term for term in dict.fromkeys(_tokens(query))
            if term not in _SEARCH_STOP_WORDS
        ]
        if not primary_terms or not self.rows:
            return []
        weighted_terms = [(term, 0.35 if term.isdigit() else 1.0) for term in primary_terms]
        related_terms = []
        for term in primary_terms:
            for root, additions in _RELATED_OFFER_TERMS.items():
                if term.startswith(root) or root.startswith(term):
                    related_terms.extend(additions)
        for term in dict.fromkeys(related_terms):
            if term not in primary_terms:
                weighted_terms.append((term, 0.3))
        total = len(self.rows)
        scored = []
        for row in self.rows:
            topical_terms = [term for term, _ in weighted_terms if not term.isdigit()]
            if topical_terms and not any(row["counts"].get(term, 0) for term in topical_terms):
                continue
            score = 0.0
            for term, term_weight in weighted_terms:
                frequency = row["counts"].get(term, 0)
                if not frequency:
                    continue
                document_frequency = self.document_frequency.get(term, 0)
                inverse_frequency = math.log(
                    1 + (total - document_frequency + 0.5) / (document_frequency + 0.5)
                )
                denominator = frequency + 1.5 * (
                    0.25 + 0.75 * row["length"] / self.average_length
                )
                score += term_weight * inverse_frequency * frequency * 2.5 / denominator
                score += term_weight * inverse_frequency * 2.0 * row["title_counts"].get(term, 0)
            if score:
                scored.append((score, row))
        scored.sort(key=lambda item: item[0], reverse=True)
        return [
            Offer(**{
                key: value for key, value in row.items()
                if key not in {"counts", "title_counts", "length"}
            })
            for _, row in scored[:limit]
        ]


class YandexSearchClient:
    def __init__(self, config: Settings, client: httpx.AsyncClient):
        self.config = config
        self.client = client

    async def search(self, query: str) -> List[Source]:
        if not self.config.yandex_search_api_key or not self.config.yandex_folder_id:
            raise ServiceConfigurationError(
                "Задайте YANDEX_SEARCH_API_KEY и YANDEX_FOLDER_ID"
            )

        payload = {
            "query": {
                "searchType": "SEARCH_TYPE_RU",
                "queryText": query,
                "familyMode": "FAMILY_MODE_MODERATE",
                "page": "0",
            },
            "sortSpec": {"sortMode": "SORT_MODE_BY_RELEVANCE"},
            "groupSpec": {
                "groupMode": "GROUP_MODE_DEEP",
                "groupsOnPage": str(self.config.search_results),
                "docsInGroup": "1",
            },
            "maxPassages": "3",
            "region": "225",
            "l10N": "LOCALIZATION_RU",
            "folderId": self.config.yandex_folder_id,
            "responseFormat": "FORMAT_XML",
        }
        try:
            response = await self.client.post(
                self.config.yandex_search_url,
                headers={"Authorization": "Api-Key " + self.config.yandex_search_api_key},
                json=payload,
            )
            response.raise_for_status()
            raw_data = response.json()["rawData"]
            xml_data = base64.b64decode(raw_data)
            return self._parse_xml(xml_data)
        except (httpx.HTTPError, KeyError, ValueError, ET.ParseError) as exc:
            raise UpstreamServiceError("Yandex Search API вернул некорректный ответ") from exc

    @staticmethod
    def _parse_xml(xml_data: bytes) -> List[Source]:
        root = ET.fromstring(xml_data)
        sources = []
        for document in root.findall(".//doc"):
            url = (document.findtext("url") or "").strip()
            parsed_url = urlparse(url)
            if parsed_url.scheme not in {"http", "https"} or not parsed_url.netloc:
                continue
            title_node = document.find("title")
            title = _plain_text("".join(title_node.itertext()) if title_node is not None else url)
            passages = [
                _plain_text("".join(node.itertext()))
                for node in document.findall(".//passage")
            ]
            snippet = " ".join(item for item in passages if item)
            if not snippet:
                snippet = _plain_text(document.findtext("headline") or "")
            sources.append(
                Source(
                    id=len(sources) + 1,
                    title=title or url,
                    url=url,
                    domain=parsed_url.netloc.removeprefix("www."),
                    snippet=snippet[:1200],
                )
            )
        return sources


class AnswerGenerator:
    def __init__(self, config: Settings, client: httpx.AsyncClient):
        self.config = config
        self.client = client

    async def generate(self, query: str, sources: List[Source]) -> str:
        if not self.config.ai_api_key or not self.config.yandex_folder_id:
            raise ServiceConfigurationError(
                "Задайте YANDEX_AI_API_KEY (или используйте YANDEX_SEARCH_API_KEY) "
                "и YANDEX_FOLDER_ID"
            )
        if not sources:
            return "По этому запросу Яндекс не нашёл подходящих источников."

        context = "\n\n".join(
            "[{id}] {title}\nURL: {url}\nФрагмент: {snippet}".format(**source.model_dump())
            for source in sources
        )
        system = (
            "Ты — исследовательский ассистент. Отвечай на языке вопроса, ясно и по делу. "
            "Используй только предоставленные результаты поиска. Каждое проверяемое утверждение "
            "сопровождай ссылкой вида [1]. Не выдумывай факты и номера источников. Если данных "
            "недостаточно, честно скажи об этом. В конце не создавай отдельный список источников."
        )
        body = {
            "model": self.config.resolved_llm_model,
            "temperature": 0.2,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": "Вопрос: " + query + "\n\nИсточники:\n" + context},
            ],
        }
        try:
            response = await self.client.post(
                self.config.llm_base_url.rstrip("/") + "/chat/completions",
                headers={
                    "Authorization": "Api-Key " + self.config.ai_api_key,
                    "Content-Type": "application/json",
                },
                content=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            )
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"].strip()
        except (httpx.HTTPError, KeyError, IndexError, ValueError) as exc:
            raise UpstreamServiceError("LLM-сервис не смог сформировать ответ") from exc
